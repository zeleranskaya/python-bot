from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook
import random

TOKEN = '1674333840:AAHRXuYwcSfKXBfu_xx1CauSFncF4u6kfvA'
book = load_workbook('база данных.xlsx')
sheet_1 = book['Лист1']
stickers_page = book['стикеры']



hi = ['приветствую, мой друг', 'я рад видеть тебя!', 'давно не виделись!']
help = ['чем тебе помочь?', 'у тебя есть какие-то проблемы?']
what = ['выбери команду, я не понимаю тебя']
questions = ['как дела?', 'как твое настроение?', 'как поживаешь?']



randomizer = random.SystemRandom()

def main():
    updater = Updater(token=TOKEN)  # объект, который ловит сообщения от Telegram

    dispather = updater.dispatcher

    handler = MessageHandler(Filters.all, do_echo)
    start_handler = CommandHandler('start', do_start)
    help_handler = CommandHandler('help', do_help)
    questions_handler = CommandHandler('questions', do_questions)
    keyboard_handler = MessageHandler(Filters.text, do_something)
    sticker_handler = MessageHandler(Filters.sticker, do_sticker)

    dispather.add_handler(sticker_handler)
    dispather.add_handler(questions_handler)
    dispather.add_handler(help_handler)
    dispather.add_handler(start_handler)
    dispather.add_handler(keyboard_handler)
    dispather.add_handler(handler)


    updater.start_polling()
    updater.idle()


def do_echo(update, context):
    keyboard = [
        ['1', '2', '3'],
        ['4', '5', '6'],
    ]

    update.message.reply_text(randomizer.choice(what),
                              reply_markup=ReplyKeyboardMarkup(
                                  keyboard, one_time_keyboard=True, resize_keyboard=True))



def do_start(update, context):

    update.message.reply_text(randomizer.choice(hi))



def do_something(update: Update, context):
    text = update.message.text

    print(stickers_page.max_row)
    for row in range(2, stickers_page.max_row + 1):
        catch_phrase = stickers_page.cell(row=row, column=4).value
        print(catch_phrase)
        print(text)
        if catch_phrase in text:
            sticker_id = stickers_page.cell(row=row, column=3).value
            print(sticker_id)
            update.message.reply_sticker(sticker_id)
            return

    if text == '1':
        update.message.reply_text('ты нажал 1', reply_markup=ReplyKeyboardRemove())
    elif text == '2':
        update.message.reply_text('ты нажал 2', reply_markup=ReplyKeyboardRemove())
    elif text == '3':
        update.message.reply_text('ты нажал 3', reply_markup=ReplyKeyboardRemove())
    else:
        do_echo(update, context)


def do_help(update: Update, context):

    user_id = update.message.from_user.id
    name = update.message.from_user.first_name
    update.message.reply_text(text=f'привет, {name}!\nтвой id: {user_id}')

def do_questions(update, context):

    update.message.reply_text(randomizer.choice(questions))

def do_sticker(update: Update, context):

    sticker_id = update.message.sticker.file_id
    update.message.reply_text(sticker_id)
    update.message.reply_sticker(sticker_id)



main()
