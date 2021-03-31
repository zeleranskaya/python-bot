from openpyxl import load_workbook


book = load_workbook('база данных.xlsx')
sheet_1 = book['Лист1']
stickers_page = book['стикеры']

print(book.worksheets)
for i in range(1, 4):
    print(stickers_page.cell(row=1, column=i).value)
