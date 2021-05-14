from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook
import random

TOKEN = '1697612810:AAFBc1j9gLPkZKBiKR0C2VK8TI4XkDVUk6Y'
book = load_workbook('база данных.xlsx')

link_page = book['ссылки']


hi = ['привет, что ты хочешь послушать сегодня?']
choice = ['выбери жанр музыки', 'какую музыку ты хочешь послушать?']
what = ['я подберу для тебя плейлист с хорошей музыкой']
mood = ['выбирай, тут много интересного']




randomizer = random.SystemRandom()

def main():
    updater = Updater(token=TOKEN)  # объект, который ловит сообщения от Telegram

    dispather = updater.dispatcher

    handler = MessageHandler(Filters.all, do_echo)
    start_handler = CommandHandler('start', do_start)
    keyboard_handler = MessageHandler(Filters.text, do_something)
    sticker_handler = MessageHandler(Filters.sticker, do_sticker)

    dispather.add_handler(sticker_handler)
    dispather.add_handler(start_handler)
    dispather.add_handler(keyboard_handler)
    dispather.add_handler(handler)


    updater.start_polling()
    print('Бот успешно запустился')
    updater.idle()


def do_echo(update, context):

    update.message.reply_text(randomizer.choice(what))



def do_start(update, context):
    keyboard = [
        ['настроение', 'жанр'],
    ]

    update.message.reply_text(randomizer.choice(hi),
                              reply_markup=ReplyKeyboardMarkup(
                                  keyboard, one_time_keyboard=True, resize_keyboard=True))


def do_something(update: Update, context):
    text = update.message.text

    print(link_page.max_row)
    for row in range(2, link_page.max_row + 1):
        catch_phrase = link_page.cell(row=row, column=1).value.strip()
        print(catch_phrase)
        print(text)
        if catch_phrase in text:
            link = link_page.cell(row=row, column=2).value
            print(link)
            update.message.reply_text(link)
            update.message.reply_text("хорошего прослушивания!")

            return do_start(update, context)

    if text == 'жанр':
        do_genre(update, context)
    elif text == 'настроение':
        do_mood(update, context)






def do_sticker(update: Update, context):

    sticker_id = update.message.sticker.file_id
    update.message.reply_text(sticker_id)
    update.message.reply_sticker(sticker_id)


def do_genre(update, context):
    keyboard = [
        ['джаз', 'рок', 'инди'],
        ['хип-хоп', 'классика', 'панк рок'],
        ['блюз', 'поп', 'метал'],
        ['hard рок', 'ретро', 'регги'],
        ['фолк', 'альтернатива', 'электроника']

    ]

    update.message.reply_text(randomizer.choice(choice),
                          reply_markup=ReplyKeyboardMarkup(
                              keyboard, one_time_keyboard=True, resize_keyboard=True))

def do_mood(update, context):
    keyboard = [
        ['вечеринка', 'хорошее настроение', 'сон'],
        ['концентрация', 'спорт', 'спокойствие'],
        ['мотивация', 'романтика', 'меланхолия']
    ]

    update.message.reply_text(randomizer.choice(mood),
                              reply_markup=ReplyKeyboardMarkup(
                                  keyboard, one_time_keyboard=True, resize_keyboard=True))

main()
